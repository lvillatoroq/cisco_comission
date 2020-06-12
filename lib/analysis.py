#!/usr/bin/python

import os, sys, re, time

#from dumper import dump
from lib.report import Report
from openpyxl import load_workbook

## Keywords used in the analysis config file

empty = 'EMPTY'
negate = 'NOT'
documentation = 'DOCUMENTATION'
always_ok = 'ALWAYS-OK'

##### ANALYZE MODULE ############

class Analysis():

    def __init__(self, filename, report_filename):
        """
        Read file with analysis_config, and initalize internal hashes and lists
        :param filename - file with analysis config
        File format:
        test name ; command ; search expression (may include defined keywords)
        'test name' will be used as key internal hashes in this object
        :return: nothing
        """

        # Switch commands, which should be searched for a specific test
        self.analysis_commands = {}

        # What to search for in a specific test
        self.analysis_searches = {}

        # Dictionary with key = IP address, value = hostname
        self.hostnames = {}

        #Dictionary with device type
        self.categories = {}

        # List used to keep the order of the anlysis from the configuration file
        self.test_names = {}

        # What will be written in the xls to dexcribe the specific test
        self.analysis_report_names = {}

        # Report object, to store test results and colors, and used to write the actual xls-file
        self.report = Report(report_filename)

        delimiter = '#####\n'
        err = ''

        workbook = load_workbook(filename)
        sheets = workbook.sheetnames
        for sheet_name in sheets:

            device_type = sheet_name
            self.categories[device_type]=[]
            self.test_names[device_type] = []
            self.analysis_report_names[device_type] = {}
            self.analysis_commands[device_type] = {}
            self.analysis_searches[device_type] = {}

            sheet = workbook[sheet_name]

            for row in sheet.rows:
                (name, command, search_string) = row[0].value, row[1].value, row[2].value

                if command == None:
                    comm = empty
                else:
                    m = re.search('^\s*((\S+)(\s+\S+)*)\s*$', command)
                    if m:
                        comm = m.group(1)
                    else:
                        comm = empty
                if search_string == None:
                    sear = empty
                else:
                    m = re.search('^\s*((\S+)(\s+\S+)*)\s*$', search_string)
                    if m:
                        sear = m.group(1)
                    else:
                        sear = empty

                self.test_names[device_type].append(name)
                self.analysis_report_names[device_type][name] = name + '\nLooking for:\n' + self.search_term(sear)
                self.analysis_commands[device_type][name] = comm
                self.analysis_searches[device_type][name] = sear

    def normalize_device_list(self, dev_list, dev_list_ok = None):
        """
        Removes devices without a known hostname
        :param dev_list - list of device IP addresses 
        :return: list of device IP addresses to use in report
        """
        if dev_list_ok is None:
            return [ d for d in dev_list if d in self.hostnames ]
        else:
            return [d for d in dev_list if d in dev_list_ok]

    def normalize_categories_list(self, cat_list, device_list):
        """
        Removes devices without a known hostname
        :param cat_list - dictionary of device addresses ith categories
        :return: list of device IP addresses to use in report in each category
        """
        category_valid = {}
        for key, value in cat_list.items():
            category_valid[key] = self.normalize_device_list(device_list,value)
        return category_valid
    
    def analyze(self, files_dir, input_files_re, device_re, hostname_re, delimiter, config_file, device_list_file):
        """
        The actual analysis of all devices
        :param files_dir - directory where output files are stored
        :param input_files_re - regexp to match for output files to use
        :param device_re - regexp to find device IP address from filename 
        :param hostname_re - regexp to find hostname from filename 
        :param delimiter - string used to delimit switch commands and output
        :param config_file - file name for configuraton of this analysis 
        :param device_list_file - file name with list of devices, used to kweep the original order
        :return: nothing
        """
        for filename in os.listdir(files_dir):
            if re.search(input_files_re, filename):
                # Store device IP address and hostname in dictionary in right category
                hostname = ''
                if re.search(hostname_re, filename):
                    m = re.search(hostname_re, filename)
                    hostname = m.group(1)
                print 'Analyzing ' + hostname
                if re.search(device_re, filename):
                    m = re.search(device_re, filename)
                    device = m.group(1)
                else:
                    device = filename

                if hostname.strip().split('-')[-1].strip()[0] == 'U':
                    category = hostname.strip().split('-')[-1].strip()[0:3]
                    if category not in self.categories:
                        category = 'Default'
                        #self.categories[category] = []
                    self.categories[category].append(device)
                else:
                    category = 'Default'
                    if 'Default' not in self.categories:
                        self.categories['Default'] = []
                    self.categories['Default'].append(device)

                self.hostnames[device] = hostname

                # Store command output in dictionary
                (command_output, err) = read_command_output(files_dir  +'/' + filename, delimiter)
                if err:
                    print err

                for key in self.analysis_commands[category]:
                    command = self.analysis_commands[category][key]
                    searching = self.analysis_searches[category][key]

                    # The actual analysis - stores results and color codes
                    found = False

                    # Empty commands, just adding lines to be popluated later on
                    if command == empty:
                        self.report.add_to_report(key, device, '', 'yellow') 
                        continue

                    # Missing command in output file
                    if not command in command_output:
                        if re.match(always_ok,searching):
                            self.report.add_to_report(key, device, 'COMMAND NOT COLLECTED\n' + command, 'green') 
                            continue
                        else:
                            self.report.add_to_report(key, device, 'COMMAND NOT COLLECTED\n' + command, 'red') 
                            continue

                    # Documantation only item
                    if re.match(documentation,searching):
                        m = re.match(documentation + '\s*(.*)',searching)
                        searching = m.group(1)
                        for l in command_output[command].split('\n'):
                            line = l.strip()
                            if re.search(searching, line):
                                self.report.add_to_report(key, device, 'FOUND\n' + line + '\n', 'white') 
                                found = True
                        if not found:
                            self.report.add_to_report(key, device, 'NOT FOUND', 'white')

                    # Item which will always be reported as OK
                    elif re.match(always_ok,searching):
                        m = re.match(always_ok + '\s*(.*)',searching)
                        searching = m.group(1)
                        for l in command_output[command].split('\n'):
                            line = l.strip()
                            if re.search(searching, line):
                                self.report.add_to_report(key, device, 'FOUND\n' + line + '\n', 'green') 
                                found = True
                        if not found:
                            self.report.add_to_report(key, device, 'NOT FOUND', 'green')

                    # Looking for output which should NOT be present
                    elif re.match(negate,searching):
                        m = re.match(negate + '\s*(.*)',searching)
                        searching = m.group(1)
                        for l in command_output[command].split('\n'):
                            line = l.strip()
                            if re.search(searching, line):
                                self.report.add_to_report(key, device, 'FOUND\n' + line + '\n', 'red') 
                                found = True
                        if not found:
                            self.report.add_to_report(key, device, 'NOT FOUND', 'green')

                    # The main case - looking for specific output in a command
                    else:
                        for l in command_output[command].split('\n'):
                            line = l.strip()
                            if re.search(searching, line):
                                self.report.add_to_report(key, device, 'FOUND\n' + line + '\n', 'green') 
                                found = True
                        if not found:
                            self.report.add_to_report(key, device, 'NOT FOUND', 'red') 
                # Note! Hirschmann sometimes returns lines with leading CRs, so we do lstrip
                #  The output key is stripped both l and r

        #Skip devices without collected commands
        #devices = self.normalize_device_list(read_devices(device_list_file))
        categories = self.normalize_categories_list(self.categories,read_devices(device_list_file))
        print(categories)
        self.report.write_xls(self.test_names, self.analysis_report_names, self.hostnames, categories)
    
    def search_term(self, text):
        """
        Building search term from the defined global keywords
        :param text - text from configurartion file, search term and optional keyword
        :return: string with regexp to use in search
        """
        searching = text
        if re.match(negate,searching):
            m = re.match(negate + '\s*(.*)',searching)
            searching = m.group(1)
        elif re.match(documentation,searching):
            m = re.match(documentation + '\s*(.*)',searching)
            searching = m.group(1)
        elif re.match(always_ok,searching):
            m = re.match(always_ok + '\s*(.*)',searching)
            searching = m.group(1)
        return searching


        
###### FUNCTIONS ###########

def read_devices(filename):
    """
    Read list of IP addresses to connect to, one address per line
    :param filename - file with list of addresses
    :return: list of devices
    """
    dev = []
    with open(filename, 'r') as lf:
        for line in lf:
            if  not re.match ('^\s*$', line) and not re.match ('^#', line):
                dev.append(line.strip())
    return dev

def read_command_output(filename, delimiter):
    """
    Read file with command output, generated by hirschmannReport.py script
    :param filename - file with list of addresses
    :param delimiter - string used to delimit command names and output
    :return: tuple:
    - dictionary with command output
    - string with errors
    """
    
    command_out = {}
    err = ''

    with open(filename, 'r') as rf:
            line = rf.readline()
            while line:
                if not re.search (delimiter, line):
                    line = rf.readline()
                    while line and not re.match (delimiter, line):
                        line = rf.readline()
                    err += '##### Internal error in file, need manual check of output #####'
                    break
            
                key = rf.readline().strip()
                line = rf.readline()
                if not re.match (delimiter, line):
                    line = rf.readline()
                    while line and not re.match (delimiter, line):
                        line = rf.readline()
                    err = err + '##### Internal error in file, need manual check of output #####'
                    break

                output = ''
                line = rf.readline()
                while line and not re.match (delimiter, line):
                    output += line
                    line = rf.readline()
                    
                command_out[key] = output

            return (command_out, err)



