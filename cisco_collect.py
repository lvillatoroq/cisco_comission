import os
import re
import threading
import openpyxl
import logging
import json
from netmiko import ConnectHandler
from datetime import datetime

#File with list of IP addresses of all devices
device_list_file = 'input/T-all.txt'

#File with credentials for accessing devices
config_file = 'input/config.txt'

#File with commands to run on devices
command_list_file = 'input/Data collection on different device typesV2.xlsx'

### This variable must match with sheet in command_list_file
command_list_device_type = 'Cisco'

### The command list must contain the line 'Commands'
### Everything in first column after that line is treated as a command and sent to the switch
command_start = 'Commands'

### This directory must exist
### Add - check for directory, if it doesnt exist, create it.
output_dir = 'files'

### General log file - summary
overall_log_file = 'files/feedback.txt'

def read_devices(filename):
    # Read list of IP addresses to connect to, one address per line
    # Parameter:
    # filename - file with list of addresses
    dev = []
    with open(filename, 'r') as lf:
        for line in lf:
            if  not re.match ('^\s*$', line) and not re.match ('^#', line):
                dev.append(line.strip())
    return dev

def read_config(filename):
    # Read config file with username and password
    # Parameter:
    # filename - file with list of addresses
    usern=''
    passw=''
    with open(filename, 'r') as lf:
        for line in lf:
            if re.match ('^\s*username\s+(\S+)', line):
                m = re.match ('^\s*username\s+(\S+)', line)
                usern = m.group(1)
            if re.match ('^\s*password\s+(\S+)', line):
                m = re.match ('^\s*password\s+(\S+)', line)
                passw = m.group(1)
    return (usern, passw)

def read_commands(filename):
    # Read commands from xls-file. Sheet name = command_list_device_type
    # Parameter:
    # filename - xls-file with list of commands
    comm = []
    workbook = openpyxl.load_workbook(filename,data_only=True)
    worksheet = workbook.get_sheet_by_name(command_list_device_type)
    #for row in worksheet.iter_rows(min_row=1, max_col=3, max_row=2):
    iter = worksheet.iter_rows()
    while True:
        row = iter.next()
        if re.search('^\s*' + command_start + '\s*$', str(row[0].value)):
            break
    for row in iter:
        command = row[0].value
        if command and not re.match ('^\s*$', command):
            comm.append(command.strip())
    return comm

#Read devices, commands and credentials from file
devices = read_devices(device_list_file)
commands = read_commands(command_list_file)
(user, pwd) = read_config(config_file)

#Add Check/Create output directory exists 
generalLogfileName = output_dir + '/' + overall_log_file
#Invalid Command Syntax
#Connects to device, sends commands and saves output to file
def ssh_to_sw (device,output_main):
    try:
        output_dict = {}
        output_dict[device]={}
        invalid_command = 'Invalid input detected at'
        #Connecting to Device
        print "connecting to: " + device
        net_connect = ConnectHandler(device_type='cisco_ios', ip=device, username=user, password=pwd)
        print "Running commands on: " + device
        #Find the hostname for device and set output file name variable
        device_name = net_connect.find_prompt()
        outputFileName = output_dir + '/' + device_name[:-1] + '_' + device +'_output.txt'
        
        with open(outputFileName, 'wb') as f:
            #set terminal length to 0 
            net_connect.send_command('terminal len 0')
            #Run commands extracted from excel file and write to output file
            for command in commands:
                f.write('###########################\r\n')
                f.write(command + '\r\n')
                f.write('###########################\r\n')
                output = net_connect.send_command(command,delay_factor=2)
                f.write(output + '\r\n\r\n')
                if 'Invalid input detected' in output:
                    output_dict[device][command] = 'Invalid Command'
            output_main.update(output_dict)
            print "Closing connection on: " + device
            #Close SSH connection to device
            net_connect.disconnect()
    except Exception as e:
            print "Failed."
            print str(e)
            output_dict[device]['exception'] = str(e)
	    output_main.update(output_dict)
  

result_dict = {}

start_time = datetime.now()
logging.basicConfig(filename='files/ssh_logs_cisco.txt', level=logging.DEBUG)
logger = logging.getLogger("netmiko")

#create threads per device
list_threads = []
for device in devices:
    t = threading.Thread(target=ssh_to_sw, args=(device,result_dict))
    t.start()
    list_threads.append(t)
    
#wait for all threads to finish 
for t in list_threads:
    t.join()

#write errors in file
with open(overall_log_file, 'wb') as file:
    file.write(json.dumps(result_dict))

end_time = datetime.now()
total_time = end_time - start_time 
print total_time










            

